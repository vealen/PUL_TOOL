import copy
import os
import shutil
import sys
import openpyxl
import pandas as pd
import pyodbc
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
from styleframe import StyleFrame, Styler, utils
from dictionaries import *



ROOT = os.path.dirname(sys.executable)


def connect_db(db_path):
    odbc_drivers = [x for x in pyodbc.drivers() if x.startswith('Microsoft Access Driver')]
    #print(odbc_drivers)
    try:
        driver = u"{Microsoft Access Driver (*.mdb, *.accdb)}"
        #print('poszedł pierwszy')
    except:
        #print('dłuższy driver nie zadziałał, lecimy dalej')
        driver = u"{Microsoft Access Driver (*.mdb)}"

    if os.path.exists(db_path):
        connected = True
        access_con_string = f"Driver={driver}; Dbq={db_path};"
        connection = pyodbc.connect(access_con_string)
        return connection, connected
    else:
        connected = False
        return 'Ścieżka nie istnieje', connected


def get_table_data(conn, sql, filter=None):
    data = pd.read_sql(sql, conn, params=filter)
    return data


def shorten_adr_for(adr_for):
    return adr_for[8:20].replace(' ', '')


def to_excel(data, excel, sheet):
    if sheet == 'LANDING_PAGE':
        pass
    else:
        with pd.ExcelWriter(excel, mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
            result = data.to_excel(writer, header=True, index=False, sheet_name=sheet, startrow=0)
            worksheet = writer.sheets[sheet] #fetch kazdy sheet -> do szerokosci kolumn
            for column_cells in worksheet.columns: #auto fit szerokosci kolumny
                length = max(len(as_text(cell.value)) for cell in column_cells)# dla kazdej kolumny max szerokosc wiersza
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length +5 #ustawienie szerokosci (z neta - dzikie akcje, do konca nie wwiem jak to dziala, jest w dokumentacji)
        return result


def as_text(value):
    if value is None:
        return ""
    return str(value)

def set_style(wb, name_sh):
    thin = Side(border_style="thin",
                color="000000")
    allign = Alignment(horizontal = 'center',
                       vertical='center',
                       wrap_text=True,
                       shrinkToFit=False)
    workbook = openpyxl.load_workbook(wb)

    sheet = workbook[name_sh] #dynamicznie przekaze w funkcji final

    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            #alignment = copy.copy(cell.alignment)
            #alignment.wrapText = True
            #cell.alignment = alignment
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = allign

    workbook.save(wb)



def set_format(wb, name_sh, headers= None):
    workbook = openpyxl.load_workbook(wb)
    sheet = workbook[name_sh]
    allign = Alignment(horizontal='right',
                       vertical='center',
                       wrap_text=True,
                       shrinkToFit=False)
    if headers:
        for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, max_col=sheet.max_column):
            if 'Jakość hodowlana' in headers:
                row[headers['Jakość hodowlana']].number_format = '00'
            if 'Zadrzewienie' in headers:
                row[headers['Zadrzewienie']].number_format = '0.00'
            if 'Pow. [ha]' in headers:
                row[headers['Pow. [ha]']].number_format = '0.00'
                row[headers['Pow. [ha]']].alignment = allign
            if 'Pow.[ha]' in headers:
                row[headers['Pow.[ha]']].number_format = '0.00'
                row[headers['Pow.[ha]']].alignment = allign
            if 'Pow. wskazania [ha]' in headers:
                row[headers['Pow. wskazania [ha]']].number_format = '0.00'
                row[headers['Pow. wskazania [ha]']].alignment = allign
            if 'Pow.[ha] ' in headers:
                row[headers['Pow.[ha] ']].number_format = '0.0000'
                row[headers['Pow.[ha] ']].alignment = allign
            if 'Pow. PNSW [ha]' in headers:
                row[headers['Pow. PNSW [ha]']].number_format = '0.00'
                row[headers['Pow. PNSW [ha]']].alignment = allign
            if '7' in headers:
                row[headers['7']].number_format  = '0.0000'
                row[headers['7']].alignment = allign
            if '9' in headers:
                row[headers['9']].number_format  = '0.0000'
                row[headers['9']].alignment = allign


    else:
        for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row,max_col=sheet.max_column, min_col=2):
                for cell in row:
                    cell.number_format = '0.00'

    workbook.save(wb)


def set_row_height(excel):
    wb = openpyxl.load_workbook(excel)
    preserve = ['LANDING_PAGE']

    for name in wb.sheetnames:
        sheet = wb[name]
        if name not in preserve:
           for row in range(4,sheet.max_row):
               sheet.row_dimensions[row].height = 10
    wb.save(excel)

def delete_empty_raports(excel, sheet_name_list):
    wb = openpyxl.load_workbook(excel)
    preserve = ['LANDING_PAGE']
    for name in wb.sheetnames:
        sheet = wb[name]
        if name not in sheet_name_list and name not in preserve:
            wb.remove_sheet(sheet)
    wb.save(excel)

def get_column_names(sheet_name,wb= None, excel = None): #zwraca nazwy kolumn (w excelu!!!) raz z ich indexem(numer kolumny ('nazwa'=index)
    if excel:
        wb = openpyxl.load_workbook(excel)
    else:
        wb = wb
    sheet = wb[sheet_name]
    headers = {}
    if sheet_name in ['Rozbieżności','ROZB']:
        for idx, col in enumerate(sheet.iter_cols(1, sheet.max_column, min_row=3), start=0):
            headers[str(col[0].value)] = idx
    else:
        for idx, col in enumerate(sheet.iter_cols(1,sheet.max_column, min_row=2),start=0):
            headers[str(col[0].value)] = idx
    return headers


def add_logo(excel, img_path, pos):
    logo = img_path
    wb = openpyxl.load_workbook(excel)
    ws = wb['LANDING_PAGE']
    img = openpyxl.drawing.image.Image(logo)
    img.anchor = pos
    ws.add_image(img)
    wb.save(excel)


def copy_blank_raport(root_dir, dst_dir, raport_name, scheme_name):  # tworzenie kopii "zapisywanie do wybranej sciezki"
    raport_name = raport_name + '.xlsx'
    scheme = os.path.join(root_dir, 'misc', scheme_name)
    new_name = os.path.join(dst_dir, raport_name)
    if not os.path.exists(new_name):
        blank_raport = shutil.copy2(scheme, dst_dir)
        os.rename(blank_raport, new_name)
        new_raport_dir = os.path.join(dst_dir, raport_name)
        return new_raport_dir
    else:
        return False

def delete_empty_raports(excel, sheet_name_list):
    wb = openpyxl.load_workbook(excel)
    preserve = ['LANDING_PAGE']
    for name in wb.sheetnames:
        sheet = wb[name]
        if name not in sheet_name_list and name not in preserve:
            wb.remove_sheet(sheet)
    wb.save(excel)


def add_logo(excel, img_path, pos):
    logo = img_path
    wb = openpyxl.load_workbook(excel)
    ws = wb['LANDING_PAGE']
    img = openpyxl.drawing.image.Image(logo)
    img.anchor = pos
    ws.add_image(img)
    wb.save(excel)


def set_bold(x):
    return 'font-weight: bold'
def set_font_size(x):
    return "font-size: 8pt "

def set_green(x):
    return 'background-color:#005040'


def set_white_text(x):
    return 'color:#ffffff'


def shorten_adr_for(adr_for):
    return adr_for[8:20].replace(' ', '')


def replace_by_dict(dict, excel_path, sheet):  # zmieniacz do strony tytulowej
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            s = ws.cell(r, c).value
            for search_txt, replace_txt in dict.items():
                if s != None and search_txt in str(s):
                    ws.cell(r, c).value = s.replace(search_txt, replace_txt)
    wb.save(excel_path)



def add_header_footer(excel_path, sheetname):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheetname]
    ws.oddHeader.center.text = "Page &[Page] of &N"
    ws.oddHeader.center.size = 8
    ws.oddHeader.center.font = "Calibri"
    ws.oddHeader.center.color = "CC3366"
    wb.save(excel_path)


def set_bold(x):
    return 'font-weight: bold'


def set_green(x):
    return 'background-color:#005040'


def set_white_text(x):
    return 'color:#ffffff'


def set_borders(dataframe):
    sf = StyleFrame(dataframe)
    sf.apply_column_style(cols_to_style=dataframe.columns,
                          styler_obj=Styler(bg_color=utils.white, font=utils.fonts.calibri, font_size=8),
                          style_header=False)
